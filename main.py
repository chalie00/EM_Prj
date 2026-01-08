# -*- coding: utf-8 -*-
import os
import random
import re
import threading
import traceback
from copy import copy
from datetime import datetime
import json
from functools import reduce
from math import gcd
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill
import tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import cv2
import easyocr
import numpy as np
import openpyxl
import pandas as pd
from tkinter import filedialog, messagebox, ttk
DENOMS = [10000, 5000, 2000, 1000]
MONTH_SHEETS = [f"{month}월" for month in range(1, 13)]
ELDER_SHEET = "어르신별 1년 사용금액"
RESTAURANT_SHEET = "식당별 월 사용금액"
COUPON_COLUMN = "쿠폰 번호"
TABLE_COLUMNS = ["날짜", "이름", "식당", "사용금액", COUPON_COLUMN, "만나이", "성별", "10000원", "5000원", "2000원", "1000원"]
AMOUNT_COLUMNS = ["10000원", "5000원", "2000원", "1000원"]
OCR_TOTAL_COLUMN = "총액"
OCR_COLUMNS = TABLE_COLUMNS + [OCR_TOTAL_COLUMN]
RESTAURANT_OPTIONS = ["왕소 숯불구이", "브레데코", "청기와 뼈해장국", "뼈집", "김밥천국", "개성진순대"]
ELDER_COLUMNS = ["이름", "월지원금", "총사용금액", "잔액"]
RESTAURANT_COLUMNS = [
    "식당명",
    "1월",
    "2월",
    "3월",
    "4월",
    "5월",
    "6월",
    "7월",
    "8월",
    "9월",
    "10월",
    "11월",
    "12월",
    "1년",
]
AMOUNT_TABLE_COLUMNS = ["No", "제품명", "단가", "수량", "금액"]
@dataclass
class CouponCell:
    image_path: Path
    grid_pos: Tuple[int, int]  # (col, row)
    coupon_number: Optional[str]
    amount: Optional[int]
    raw_texts: List[str]

def normalize_coupon_key(value: object) -> str:
    raw = re.sub(r"\s+", "", str(value or "")).strip()
    digits = re.sub(r"[^0-9]", "", raw)
    if digits == "":
        return ""
    return digits.lstrip("0") or "0"
class OCRProcessor:
    def __init__(self) -> None:
        # easyocr handles Korean + English without needing a local Tesseract binary.
        self.reader = easyocr.Reader(["ko", "en"], gpu=False, verbose=False)
    def split_grid(self, image: np.ndarray, cols: int, rows: int) -> List[Tuple[np.ndarray, Tuple[int, int]]]:
        h, w = image.shape[:2]
        cell_w, cell_h = w // cols, h // rows
        cells = []
        for r in range(rows):
            for c in range(cols):
                x0, y0 = c * cell_w, r * cell_h
                cell = image[y0 : y0 + cell_h, x0 : x0 + cell_w]
                cells.append((cell, (c, r)))
        return cells
    def parse_amount(self, detections: List[Tuple]) -> Optional[int]:
        candidates = []
        for _, text, conf in detections:
            cleaned = text.replace(",", "").replace(" ", "")
            digits = re.sub(r"[^0-9]", "", cleaned)
            if not digits:
                continue
            try:
                value = int(digits)
            except ValueError:
                continue
            candidates.append((value, conf))
        if not candidates:
            return None
        exact_matches = [(value, conf) for value, conf in candidates if value in DENOMS]
        if exact_matches:
            return max(exact_matches, key=lambda x: x[1])[0]
        # choose the candidate that matches closest to allowed denominations
        best = None
        best_score = -1
        for value, conf in candidates:
            closest = min(DENOMS, key=lambda v: abs(v - value))
            score = conf - (abs(closest - value) / 10000)  # small penalty for distance
            if score > best_score:
                best_score = score
                best = closest
        return best
    def parse_coupon_number(self, detections: List[Tuple]) -> Optional[str]:
        label_hits: List[Tuple[str, float]] = []
        numeric_hits: List[Tuple[str, float, float]] = []  # text, conf, vertical_pos
        for bbox, text, conf in detections:
            cleaned = text.strip()
            digits_only = re.sub(r"[^0-9]", "", cleaned)
            if "no" in cleaned.lower():
                if digits_only:
                    label_hits.append((digits_only, conf))
            if digits_only:
                y_positions = [p[1] for p in bbox]
                avg_y = sum(y_positions) / len(y_positions)
                numeric_hits.append((digits_only, conf, avg_y))
        if label_hits:
            return max(label_hits, key=lambda x: x[1])[0] or None
        if numeric_hits:
            # prefer the lowest number on the ticket (likely near the bottom)
            numeric_hits.sort(key=lambda x: (x[2], -x[1]), reverse=True)
            return numeric_hits[0][0] or None
        return None
    def ocr_cell(self, cell_image: np.ndarray) -> Tuple[Optional[int], Optional[str], List[str]]:
        detections = self.reader.readtext(cell_image)
        amount = self.parse_amount(detections)
        number = self.parse_coupon_number(detections)
        if amount is None or amount == 1000:
            amount = self.ocr_amount_focus(cell_image) or amount
        raw = [t for _, t, _ in detections]
        return amount, number, raw
    def ocr_amount_focus(self, cell_image: np.ndarray) -> Optional[int]:
        h, w = cell_image.shape[:2]
        top = cell_image[: int(h * 0.6), :]
        gray = cv2.cvtColor(top, cv2.COLOR_BGR2GRAY)
        resized = cv2.resize(gray, (w * 2, int(h * 0.6) * 2), interpolation=cv2.INTER_CUBIC)
        _, thresh = cv2.threshold(resized, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        detections = self.reader.readtext(thresh, allowlist="0123456789")
        return self.parse_amount(detections)
    def process_image(self, image_path: Path, cols: int, rows: int) -> List[CouponCell]:
        data = np.fromfile(str(image_path), dtype=np.uint8)
        image = cv2.imdecode(data, cv2.IMREAD_COLOR)
        if image is None:
            raise RuntimeError(f"이미지를 불러올 수 없습니다: {image_path}")
        cells = self.split_grid(image, cols, rows)
        results: List[CouponCell] = []
        for cell_img, pos in cells:
            amount, number, raw = self.ocr_cell(cell_img)
            results.append(CouponCell(image_path=image_path, grid_pos=pos, coupon_number=number, amount=amount, raw_texts=raw))
        return results
def normalize_import_df(df: pd.DataFrame, columns: List[str], header_labels: List[str]) -> pd.DataFrame:
    def norm(value: str) -> str:
        return re.sub(r"\s+", "", str(value or "")).strip()
    header_row = None
    best_match = 0
    norm_labels = {norm(label) for label in header_labels if label}
    min_hit = min(len(norm_labels), 4) if norm_labels else 1
    for idx, row in df.iterrows():
        row_values = [norm(v) for v in row.tolist()]
        matches = sum(1 for v in row_values if v in norm_labels)
        if matches > best_match:
            best_match = matches
            header_row = idx
        if matches >= max(2, min_hit):
            break
    if header_row is None or best_match == 0:
        raise ValueError("헤더를 찾을 수 없습니다.")
    header = [str(v).strip() if pd.notna(v) else "" for v in df.iloc[header_row].tolist()]
    # Fill missing header names with expected columns in order, when possible.
    if columns:
        missing_cols = [col for col in columns if col not in header]
        if missing_cols:
            filled_header = []
            missing_iter = iter(missing_cols)
            for name in header:
                if name:
                    filled_header.append(name)
                else:
                    next_name = next(missing_iter, "")
                    filled_header.append(next_name)
            header = filled_header
    data = df.iloc[header_row + 1 :].reset_index(drop=True)
    data.columns = header
    rename_map = {}
    if COUPON_COLUMN not in data.columns and "쿠폰번호" in data.columns:
        rename_map["쿠폰번호"] = COUPON_COLUMN
    legacy_amounts = {
        "10000원권": "10000원",
        "5000원권": "5000원",
        "2000원권": "2000원",
        "1000원권": "1000원",
    }
    for old, new in legacy_amounts.items():
        if new not in data.columns and old in data.columns:
            rename_map[old] = new
    if rename_map:
        data = data.rename(columns=rename_map)
    for col in columns:
        if col not in data.columns:
            data[col] = None
    data = data[columns]
    if COUPON_COLUMN in data.columns:
        def format_coupon(value: object) -> object:
            if pd.isna(value):
                return value
            if isinstance(value, (int, np.integer)):
                return str(int(value)).zfill(3)
            if isinstance(value, (float, np.floating)):
                if float(value).is_integer():
                    return str(int(value)).zfill(3)
            raw = str(value).strip()
            digits = re.sub(r"[^0-9]", "", raw)
            if digits == "":
                return raw
            return digits.zfill(3)
        data[COUPON_COLUMN] = data[COUPON_COLUMN].apply(format_coupon)
    for col in AMOUNT_COLUMNS:
        if col in data.columns:
            data[col] = pd.to_numeric(data[col], errors="coerce").fillna(0).astype(int)
    if "사용금액" in data.columns:
        data["사용금액"] = pd.to_numeric(data["사용금액"], errors="coerce").fillna(0).astype(int)
    if "월지원금" in data.columns:
        data["월지원금"] = pd.to_numeric(data["월지원금"], errors="coerce").fillna(0).astype(int)
    if "총사용금액" in data.columns:
        data["총사용금액"] = pd.to_numeric(data["총사용금액"], errors="coerce").fillna(0).astype(int)
    if "잔액" in data.columns:
        data["잔액"] = pd.to_numeric(data["잔액"], errors="coerce").fillna(0).astype(int)
    return data
def find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, header_labels: List[str]) -> int:
    def norm(value: str) -> str:
        return re.sub(r"\s+", "", str(value or "")).strip()
    norm_labels = {norm(label) for label in header_labels if label}
    best_row = None
    best_match = 0
    min_hit = min(len(norm_labels), 4) if norm_labels else 1
    for row in ws.iter_rows():
        row_values = [norm(cell.value) for cell in row]
        matches = sum(1 for v in row_values if v in norm_labels)
        if matches > best_match:
            best_match = matches
            best_row = row[0].row if row else None
        if matches >= max(2, min_hit):
            break
    if best_row is None or best_match == 0:
        raise ValueError("헤더를 찾을 수 없습니다.")
    return best_row

def get_sheet_columns(sheet_name: str) -> List[str]:
    if sheet_name == ELDER_SHEET:
        return ELDER_COLUMNS
    if sheet_name == RESTAURANT_SHEET:
        return RESTAURANT_COLUMNS
    return TABLE_COLUMNS

def get_sheet_header_labels(sheet_name: str) -> List[str]:
    if sheet_name == ELDER_SHEET:
        return ["이름", "월지원금", "총사용금액", "잔액"]
    if sheet_name == RESTAURANT_SHEET:
        return ["식당명", "1월", "2월", "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월", "1년"]
    return TABLE_COLUMNS + ["쿠폰번호", "10000원권", "5000원권", "2000원권", "1000원권"]
def build_export_workbook(
    df: pd.DataFrame,
    path: Path,
    template_path: Optional[Path] = None,
    original_df: Optional[pd.DataFrame] = None,
    sheet_name: Optional[str] = None,
) -> None:
    target_sheet = sheet_name or (MONTH_SHEETS[0] if MONTH_SHEETS else "")
    columns = get_sheet_columns(target_sheet)
    def is_highlight_fill(cell: openpyxl.cell.cell.Cell) -> bool:
        try:
            return cell.fill is not None and cell.fill.start_color is not None and cell.fill.start_color.value == "FFFF00"
        except Exception:
            return False
    def update_restaurant_sheet(wb: openpyxl.Workbook) -> None:
        restaurants: List[str] = []
        for month in MONTH_SHEETS:
            if month not in wb.sheetnames:
                continue
            ws_month = wb[month]
            try:
                header_row = find_header_row(ws_month, get_sheet_header_labels(month))
            except ValueError:
                continue
            headers = [str(ws_month.cell(header_row, c).value or "").strip() for c in range(1, ws_month.max_column + 1)]
            if "식당" not in headers:
                continue
            rest_col = headers.index("식당") + 1
            for r in range(header_row + 1, ws_month.max_row + 1):
                value = ws_month.cell(r, rest_col).value
                name = str(value or "").strip()
                if name and name not in restaurants:
                    restaurants.append(name)
        if RESTAURANT_SHEET in wb.sheetnames:
            ws = wb[RESTAURANT_SHEET]
        else:
            ws = wb.create_sheet(title=RESTAURANT_SHEET)
        try:
            header_row = find_header_row(ws, get_sheet_header_labels(RESTAURANT_SHEET))
        except ValueError:
            header_row = 1
            for col_idx, col_name in enumerate(RESTAURANT_COLUMNS, 1):
                ws.cell(row=header_row, column=col_idx, value=col_name)
        data_start = header_row + 1
        template_row = data_start
        for idx, name in enumerate(restaurants, start=data_start):
            for col_idx, col_name in enumerate(RESTAURANT_COLUMNS, 1):
                src_cell = ws.cell(row=template_row, column=col_idx)
                dest_cell = ws.cell(row=idx, column=col_idx)
                if idx != template_row:
                    dest_cell._style = copy(src_cell._style)
                src_value = src_cell.value
                if src_value is not None and (src_cell.data_type == "f" or str(src_value).startswith("=")):
                    if idx == template_row:
                        dest_cell.value = src_value
                    else:
                        try:
                            dest_cell.value = Translator(str(src_value), origin=src_cell.coordinate).translate_formula(
                                dest_cell.coordinate
                            )
                        except Exception:
                            dest_cell.value = str(src_value)
                    continue
                if col_idx == 1:
                    dest_cell.value = name
                elif 2 <= col_idx <= 13:
                    month = MONTH_SHEETS[col_idx - 2]
                    dest_cell.value = f"=SUMIF('{month}'!C:C, $A{idx}, '{month}'!D:D)"
                elif col_name == "1년":
                    dest_cell.value = f"=SUM(B{idx}:M{idx})"
        last_row = data_start + len(restaurants) - 1
        for r in range(last_row + 1, ws.max_row + 1):
            for c in range(1, len(RESTAURANT_COLUMNS) + 1):
                ws.cell(row=r, column=c).value = None

    if template_path and template_path.exists():
        wb = openpyxl.load_workbook(template_path)
        ws = wb[target_sheet] if target_sheet in wb.sheetnames else wb.active
        header_row = find_header_row(ws, get_sheet_header_labels(target_sheet))
        data_start = header_row + 1
        template_row = data_start
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        original_map: Dict[str, Dict[str, int]] = {}
        if original_df is not None and not original_df.empty and COUPON_COLUMN in columns:
            for _, row in original_df.iterrows():
                coupon_no = str(row.get(COUPON_COLUMN, "")).strip()
                if not coupon_no:
                    continue
                original_map[coupon_no] = {col: int(row.get(col, 0)) for col in AMOUNT_COLUMNS}
        for idx, (_, row) in enumerate(df.iterrows()):
            row_idx = data_start + idx
            if row_idx > ws.max_row:
                ws.append([None] * len(columns))
            for col_idx, col_name in enumerate(columns, 1):
                src_cell = ws.cell(row=template_row, column=col_idx)
                dest_cell = ws.cell(row=row_idx, column=col_idx)
                if row_idx != template_row:
                    dest_cell._style = copy(src_cell._style)
                src_value = src_cell.value
                if src_value is not None and (src_cell.data_type == "f" or str(src_value).startswith("=")):
                    if row_idx == template_row:
                        dest_cell.value = src_value
                    else:
                        try:
                            dest_cell.value = Translator(str(src_value), origin=src_cell.coordinate).translate_formula(
                                dest_cell.coordinate
                            )
                        except Exception:
                            dest_cell.value = str(src_value)
                else:
                    dest_cell.value = row.get(col_name, None)
                highlight_restaurant = col_name == "식당" and int(row.get("__restaurant_filled") or 0) == 1
                if col_name == "식당":
                    if highlight_restaurant:
                        dest_cell.fill = highlight_fill
                    else:
                        dest_cell.fill = PatternFill(fill_type=None)
                if col_name in AMOUNT_COLUMNS and COUPON_COLUMN in columns:
                    if int(row.get("__amounts_changed") or 0) != 1:
                        dest_cell.fill = PatternFill(fill_type=None)
                        continue
                    coupon_no = str(row.get(COUPON_COLUMN, "")).strip()
                    original_vals = original_map.get(coupon_no, {})
                    original_value = int(original_vals.get(col_name, 0))
                    new_value = int(row.get(col_name, 0))
                    if new_value != original_value:
                        dest_cell.fill = highlight_fill
                    else:
                        dest_cell.fill = PatternFill(fill_type=None)
        # clear extra old rows
        last_row = data_start + len(df) - 1
        for row_idx in range(last_row + 1, ws.max_row + 1):
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).value = None
        update_restaurant_sheet(wb)
        wb.save(path)
        return
    wb = openpyxl.Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    def write_month_sheet(sheet: str, sheet_df: pd.DataFrame) -> None:
        ws = wb.create_sheet(title=sheet)
        for col_idx, col_name in enumerate(TABLE_COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for idx, (_, row) in enumerate(sheet_df.iterrows(), start=2):
            for col_idx, col_name in enumerate(TABLE_COLUMNS, 1):
                if col_name == "사용금액":
                    h_col = openpyxl.utils.get_column_letter(TABLE_COLUMNS.index("10000원") + 1)
                    i_col = openpyxl.utils.get_column_letter(TABLE_COLUMNS.index("5000원") + 1)
                    j_col = openpyxl.utils.get_column_letter(TABLE_COLUMNS.index("2000원") + 1)
                    k_col = openpyxl.utils.get_column_letter(TABLE_COLUMNS.index("1000원") + 1)
                    ws.cell(
                        row=idx,
                        column=col_idx,
                        value=f"={h_col}{idx}*10000+{i_col}{idx}*5000+{j_col}{idx}*2000+{k_col}{idx}*1000",
                    )
                else:
                    ws.cell(row=idx, column=col_idx, value=row.get(col_name, None))

    def write_elder_sheet(names: List[str]) -> None:
        ws = wb.create_sheet(title=ELDER_SHEET)
        for col_idx, col_name in enumerate(ELDER_COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for idx, name in enumerate(names, start=2):
            ws.cell(row=idx, column=1, value=name)
            ws.cell(row=idx, column=2, value=300000)
            sum_parts = [f"SUMIF('{month}'!B:B, $A{idx}, '{month}'!D:D)" for month in MONTH_SHEETS]
            ws.cell(row=idx, column=3, value=f"={'+'.join(sum_parts)}")
            ws.cell(row=idx, column=4, value=f"=B{idx}-C{idx}")

    def write_restaurant_sheet(restaurants: List[str]) -> None:
        ws = wb.create_sheet(title=RESTAURANT_SHEET)
        for col_idx, col_name in enumerate(RESTAURANT_COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for idx, name in enumerate(restaurants, start=2):
            ws.cell(row=idx, column=1, value=name)
            for month_idx, month in enumerate(MONTH_SHEETS, start=2):
                ws.cell(row=idx, column=month_idx, value=f"=SUMIF('{month}'!C:C, $A{idx}, '{month}'!D:D)")
            ws.cell(row=idx, column=len(RESTAURANT_COLUMNS), value=f"=SUM(B{idx}:M{idx})")

    month_df = df if target_sheet in MONTH_SHEETS else pd.DataFrame(columns=TABLE_COLUMNS)
    for month in MONTH_SHEETS:
        write_month_sheet(month, month_df if month == target_sheet else pd.DataFrame(columns=TABLE_COLUMNS))
    names = sorted(set(str(v).strip() for v in month_df.get("이름", pd.Series([], dtype=str)) if str(v).strip()))
    restaurants = sorted(set(str(v).strip() for v in month_df.get("식당", pd.Series([], dtype=str)) if str(v).strip()))
    write_elder_sheet(names)
    write_restaurant_sheet(restaurants)
    wb.save(path)
class App:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Coupon OCR & Excel Merger")
        try:
            self.root.state("zoomed")  # windows full screen
        except tk.TclError:
            self.root.attributes("-zoomed", True)
        self.root.minsize(1100, 700)
        self.processor = OCRProcessor()
        self.import_df: Optional[pd.DataFrame] = None
        self.import_sheet: Optional[str] = None
        self.ocr_dfs: Dict[str, pd.DataFrame] = {}
        self.ocr_df = pd.DataFrame(columns=TABLE_COLUMNS)
        self.participant_map: Dict[str, Dict[str, str]] = {}
        self.participant_df: Optional[pd.DataFrame] = None
        self.sheet_names: List[str] = []
        self.sheet_var = tk.StringVar(value="")
        self.restaurant_input_var = tk.StringVar(value="")
        self.restaurant_options = RESTAURANT_OPTIONS.copy()
        self._load_restaurant_options()
        self.photo_files: List[Path] = []
        self._export_in_progress = False
        self._edit_entry: Optional[tk.Entry] = None
        self._last_export_path: Optional[Path] = None
        self._import_path: Optional[Path] = None
        self.amount_df = pd.DataFrame(columns=AMOUNT_TABLE_COLUMNS)
        self.amount_target_var = tk.StringVar(value="")
        self._load_last_export_path()
        self._build_ui()
        sample_excel = Path("Sample/Sample_Excel.xlsx")
        if sample_excel.exists():
            try:
                self.load_excel(sample_excel)
            except Exception:
                pass
    def _build_ui(self) -> None:
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        top_bar = ttk.Frame(self.root, padding=10)
        top_bar.grid(row=0, column=0, sticky="ew")
        top_bar.columnconfigure(2, weight=1)

        source_frame = ttk.LabelFrame(top_bar, text="원본/사진 분석", padding=8)
        source_frame.grid(row=0, column=0, padx=(0, 8), sticky="w")

        ttk.Button(source_frame, text="원본 불러오기", command=self.browse_excel).grid(row=0, column=0, padx=4, pady=2)
        ttk.Button(source_frame, text="참여자 불러오기", command=self.browse_participant).grid(
            row=0, column=1, padx=4, pady=2
        )
        ttk.Label(source_frame, text="Sheet:").grid(row=0, column=2, padx=(12, 2), pady=2)
        self.sheet_combo = ttk.Combobox(source_frame, textvariable=self.sheet_var, state="disabled", width=18)
        self.sheet_combo.grid(row=0, column=3, padx=4, pady=2)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)
        ttk.Button(source_frame, text="사진 선택 (복수)", command=self.browse_photos).grid(row=0, column=4, padx=4, pady=2)
        ttk.Label(source_frame, text="Grid:").grid(row=0, column=5, padx=(12, 2), pady=2)
        self.cols_var = tk.IntVar(value=2)
        self.rows_var = tk.IntVar(value=3)
        ttk.Spinbox(source_frame, from_=1, to=5, textvariable=self.cols_var, width=4).grid(row=0, column=6, pady=2)
        ttk.Label(source_frame, text="x").grid(row=0, column=7, pady=2)
        ttk.Spinbox(source_frame, from_=1, to=5, textvariable=self.rows_var, width=4).grid(row=0, column=8, pady=2)
        ttk.Button(source_frame, text="사진 분석", command=self.run_analysis).grid(row=0, column=9, padx=8, pady=2)
        ttk.Button(source_frame, text="내보내기", command=self.export_excel).grid(row=0, column=10, padx=4, pady=2)
        ttk.Button(source_frame, text="폴더 열기", command=self.open_export_folder).grid(row=0, column=11, padx=4, pady=2)

        amount_frame = ttk.LabelFrame(top_bar, text="금액 계산", padding=8)
        amount_frame.grid(row=0, column=1, padx=(0, 8), sticky="w")

        ttk.Button(amount_frame, text="금액 계산 원본 열기", command=self.open_amount_source).grid(row=0, column=0, padx=4, pady=2)
        ttk.Label(amount_frame, text="원하는 금액:").grid(row=0, column=1, padx=(12, 2), pady=2)
        ttk.Entry(amount_frame, textvariable=self.amount_target_var, width=10).grid(row=0, column=2, pady=2)
        ttk.Button(amount_frame, text="금액수정하기", command=self.adjust_amounts).grid(row=0, column=3, padx=8, pady=2)
        ttk.Button(amount_frame, text="금액 내보내기", command=self.export_amount_table).grid(row=0, column=4, padx=4, pady=2)
        ttk.Label(amount_frame, text="식당 이름:").grid(row=0, column=5, padx=(12, 2), pady=2)
        ttk.Entry(amount_frame, textvariable=self.restaurant_input_var, width=16).grid(row=0, column=6, pady=2)
        ttk.Button(amount_frame, text="식당저장", command=self.save_restaurant).grid(row=0, column=7, padx=4, pady=2)
        ttk.Button(amount_frame, text="식당삭제", command=self.delete_restaurant).grid(row=0, column=8, padx=4, pady=2)

        self.status_var = tk.StringVar(value="준비 완료")
        ttk.Label(top_bar, textvariable=self.status_var).grid(row=0, column=2, sticky="e")

        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.grid(row=1, column=0, sticky="nsew")
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=0)
        main_frame.columnconfigure(2, weight=1)
        main_frame.columnconfigure(3, weight=0)
        main_frame.rowconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)
        main_frame.rowconfigure(5, weight=1)

        # Imported excel panel
        self.import_label = ttk.Label(main_frame, text="원본 파일")
        self.import_label.grid(row=0, column=0, columnspan=2, sticky="w")
        self.import_tree, import_vsb = self._make_tree(main_frame)
        self.import_tree.grid(row=1, column=0, sticky="nsew", padx=(0, 2))
        import_vsb.grid(row=1, column=1, sticky="ns")
        self.import_tree.tag_configure("hover", background="#dff0d8")
        self.import_tree.bind("<Motion>", self.on_import_hover)
        self.import_tree.bind("<Leave>", self.on_import_leave)

        ttk.Label(main_frame, text="참여자 목록").grid(row=2, column=0, columnspan=2, sticky="w", pady=(8, 0))
        self.participant_tree = ttk.Treeview(main_frame, columns=[], show="headings", height=8)
        participant_vsb = ttk.Scrollbar(main_frame, orient="vertical", command=self.participant_tree.yview)
        self.participant_tree.configure(yscrollcommand=participant_vsb.set)
        self.participant_tree.grid(row=3, column=0, sticky="nsew", padx=(0, 2))
        participant_vsb.grid(row=3, column=1, sticky="ns")

        # OCR result panel
        ttk.Label(main_frame, text="사진 분석 결과").grid(row=0, column=2, columnspan=2, sticky="w")
        self.ocr_tree, ocr_vsb = self._make_tree(main_frame)
        self.ocr_tree.grid(row=1, column=2, sticky="nsew", padx=(2, 0))
        ocr_vsb.grid(row=1, column=3, sticky="ns")
        self.ocr_tree.bind("<Double-1>", self.on_ocr_double_click)
        self.ocr_tree.bind("<ButtonRelease-1>", self.on_ocr_single_click)
        self.ocr_tree.bind("<Button-3>", self.on_ocr_right_click)
        self.ocr_menu = tk.Menu(self.root, tearoff=0)
        self.ocr_menu.add_command(label="삭제", command=self.delete_selected_ocr)
        self._populate_tree(self.ocr_tree, self.ocr_df)

        ttk.Label(main_frame, text="금액 계산 목록").grid(row=2, column=2, columnspan=2, sticky="w", pady=(8, 0))
        self.amount_tree, amount_vsb = self._make_amount_tree(main_frame)
        self.amount_tree.grid(row=3, column=2, sticky="nsew", padx=(2, 0))
        amount_vsb.grid(row=3, column=3, sticky="ns")
        self.amount_tree.bind("<Double-1>", self.on_amount_double_click)
        self._populate_amount_tree()

        # Missing info panel
        missing_frame = ttk.Frame(self.root, padding=10)
        missing_frame.grid(row=2, column=0, sticky="ew")
        ttk.Label(missing_frame, text="원본 파일에 없는 쿠폰 번호 (이미지 파일명, 위치)").grid(row=0, column=0, sticky="w")
        self.missing_text = tk.Text(missing_frame, height=4)
        self.missing_text.grid(row=1, column=0, sticky="ew")
        missing_frame.columnconfigure(0, weight=1)
    def _configure_participant_tree(self, columns: List[str]) -> None:
        self.participant_tree["columns"] = columns
        for col in columns:
            self.participant_tree.heading(col, text=col)
            self.participant_tree.column(col, width=self._get_column_width(col), anchor="center")

    def _populate_participant_tree(self) -> None:
        for item in self.participant_tree.get_children():
            self.participant_tree.delete(item)
        if self.participant_df is None or self.participant_df.empty:
            return
        columns = list(self.participant_tree["columns"])
        for _, row in self.participant_df.iterrows():
            self.participant_tree.insert("", "end", values=[row.get(col, "") for col in columns])
    def _get_column_width(self, column: str) -> int:
        widths = {
            "날짜": 90,
            "이름": 80,
            "식당": 140,
            "사용금액": 90,
            COUPON_COLUMN: 90,
            "만나이": 60,
            "성별": 60,
            "10000원": 70,
            "5000원": 70,
            "2000원": 70,
            "1000원": 70,
            "월지원금": 90,
            "총사용금액": 90,
            "잔액": 90,
            "식당명": 140,
            "1년": 80,
        }
        if column.endswith("월") and column[:-1].isdigit():
            return 70
        return widths.get(column, 90)

    def _configure_tree(self, tree: ttk.Treeview, columns: List[str]) -> None:
        tree["columns"] = columns
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=self._get_column_width(col), anchor="center")

    def _make_tree(self, parent: tk.Widget) -> Tuple[ttk.Treeview, ttk.Scrollbar]:
        tree = ttk.Treeview(parent, columns=TABLE_COLUMNS, show="headings", height=12)
        self._configure_tree(tree, TABLE_COLUMNS)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        return tree, vsb
    def _make_amount_tree(self, parent: tk.Widget) -> Tuple[ttk.Treeview, ttk.Scrollbar]:
        tree = ttk.Treeview(parent, columns=AMOUNT_TABLE_COLUMNS, show="headings", height=8, selectmode="browse")
        widths = {"No": 25, "제품명": 160, "단가": 80, "수량": 70, "금액": 90}
        for col in AMOUNT_TABLE_COLUMNS:
            tree.heading(col, text=col)
            tree.column(col, width=widths.get(col, 90), anchor="center")
        vsb = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        return tree, vsb
    def _refresh_amount_numbers(self) -> None:
        if self.amount_df is None:
            return
        self.amount_df = self.amount_df.reset_index(drop=True)
        self.amount_df["No"] = [idx + 1 for idx in range(len(self.amount_df))]
        if "단가" in self.amount_df.columns and "수량" in self.amount_df.columns:
            self.amount_df["금액"] = (self.amount_df["단가"].fillna(0).astype(int)
                                    * self.amount_df["수량"].fillna(0).astype(int))
    def _populate_amount_tree(self) -> None:
        for item in self.amount_tree.get_children():
            self.amount_tree.delete(item)
        if self.amount_df is not None and not self.amount_df.empty:
            self._refresh_amount_numbers()
            for _, row in self.amount_df.iterrows():
                values = [row.get(col, "") for col in AMOUNT_TABLE_COLUMNS]
                self.amount_tree.insert("", "end", values=values)
        total = 0
        if self.amount_df is not None and not self.amount_df.empty:
            total = int(self.amount_df.get("금액", pd.Series([0])).fillna(0).astype(int).sum())
        total_row = [""] * len(AMOUNT_TABLE_COLUMNS)
        total_row[2] = "Total"
        total_row[-1] = str(total)
        self.amount_tree.insert("", "end", values=total_row)
        self.amount_tree.insert("", "end", values=[""] * len(AMOUNT_TABLE_COLUMNS))
    def _populate_tree(self, tree: ttk.Treeview, df: pd.DataFrame) -> None:
        for item in tree.get_children():
            tree.delete(item)
        columns = list(tree["columns"])
        for _, row in df.iterrows():
            tree.insert("", "end", values=[row.get(col, "") for col in columns])
        if tree is self.ocr_tree:
            tree.insert("", "end", values=[""] * len(columns))
    def browse_photos(self) -> None:
        files = filedialog.askopenfilenames(
            title="쿠폰 사진 선택",
            filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.tif"), ("All files", "*.*")],
        )
        if files:
            self.photo_files = [Path(f) for f in files]
            self.status_var.set(f"{len(self.photo_files)}장 선택됨")
    def browse_excel(self) -> None:
        file = filedialog.askopenfilename(title="Excel 파일 선택", filetypes=[("Excel", "*.xlsx;*.xls")])
        if file:
            self.load_excel(Path(file))
    def browse_participant(self) -> None:
        file = filedialog.askopenfilename(title="참여자 파일 선택", filetypes=[("Excel", "*.xlsx;*.xls")])
        if file:
            self.load_participant_map(Path(file))

    def _load_last_export_path(self) -> None:
        path = Path("last_export_path.txt")
        if not path.exists():
            return
        try:
            saved = path.read_text(encoding="utf-8").strip()
        except OSError:
            return
        if saved:
            self._last_export_path = Path(saved)

    def _save_last_export_path(self) -> None:
        if not self._last_export_path:
            return
        try:
            Path("last_export_path.txt").write_text(str(self._last_export_path), encoding="utf-8")
        except OSError:
            pass

    def open_amount_source(self) -> None:
        file = filedialog.askopenfilename(title="금액 계산 원본 선택", filetypes=[("Excel", "*.xlsx;*.xls")])
        if file:
            self.load_amount_excel(Path(file))

    def save_restaurant(self) -> None:
        name = self.restaurant_input_var.get().strip()
        if not name:
            messagebox.showwarning("경고", "식당 이름을 입력하세요.")
            return
        if name not in self.restaurant_options:
            self.restaurant_options.append(name)
            self._save_restaurant_options()
        self.restaurant_input_var.set("")

    def delete_restaurant(self) -> None:
        name = self.restaurant_input_var.get().strip()
        if not name:
            messagebox.showwarning("경고", "식당 이름을 입력하세요.")
            return
        if name in self.restaurant_options:
            self.restaurant_options.remove(name)
            self._save_restaurant_options()
        self.restaurant_input_var.set("")

    def _restaurant_options_path(self) -> Path:
        return Path("restaurant_options.json")

    def _load_restaurant_options(self) -> None:
        path = self._restaurant_options_path()
        if not path.exists():
            return
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return
        if isinstance(data, list) and data:
            self.restaurant_options = [str(item) for item in data if str(item).strip()]

    def _save_restaurant_options(self) -> None:
        path = self._restaurant_options_path()
        try:
            path.write_text(json.dumps(self.restaurant_options, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass

    def load_amount_excel(self, path: Path) -> None:
        df = pd.read_excel(path)
        df.columns = [str(col).strip() for col in df.columns]
        if "제품명" not in df.columns or "단가" not in df.columns:
            messagebox.showwarning("경고", "금액 계산 원본에는 '제품명', '단가' 컬럼이 필요합니다.")
            return
        work = df.copy()
        work["제품명"] = work["제품명"].fillna("").astype(str).str.strip()
        work["단가"] = pd.to_numeric(work["단가"], errors="coerce").fillna(0).astype(int)
        if "수량" in work.columns:
            work["수량"] = pd.to_numeric(work["수량"], errors="coerce").fillna(0).astype(int)
        else:
            work["수량"] = 0

        work = work[(work["제품명"] != "") | (work["단가"] > 0) | (work["수량"] > 0)]
        rows: List[Dict[str, object]] = []
        for _, row in work.iterrows():
            rows.append(
                {
                    "No": len(rows) + 1,
                    "제품명": row.get("제품명", ""),
                    "단가": int(row.get("단가", 0)),
                    "수량": int(row.get("수량", 0)),
                    "금액": 0,
                }
            )
        self.amount_df = pd.DataFrame(rows, columns=AMOUNT_TABLE_COLUMNS)
        self._populate_amount_tree()
        self.status_var.set(f"금액 계산 원본 로드: {path.name}")
    def load_excel(self, path: Path) -> None:
        sheet_names: List[str] = []
        try:
            excel_file = pd.ExcelFile(path)
            sheet_names = excel_file.sheet_names
        except Exception:
            try:
                wb = openpyxl.load_workbook(path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            except Exception:
                sheet_names = []
        known_sheets = MONTH_SHEETS + [ELDER_SHEET, RESTAURANT_SHEET]
        self.sheet_names = [name for name in sheet_names if name in known_sheets] or sheet_names
        self._import_path = path
        self.ocr_dfs = {}
        self.import_df = None
        self.import_sheet = None
        default_sheet = self.sheet_names[0] if self.sheet_names else ""
        self.sheet_var.set(default_sheet)
        self._update_sheet_combo()
        if default_sheet:
            self._load_sheet_data(default_sheet)
        else:
            self.status_var.set(f"Excel 로드 완료: {path.name}")

    def load_participant_map(self, path: Path) -> None:
        try:
            df_raw = pd.read_excel(path, header=None)
        except Exception as exc:
            messagebox.showerror("오류", f"참여자 파일 로드 실패: {exc}")
            return
        header_labels = ["연번", "성명", "이름", "성별", "상담여부", "상담 여부", "상담"]
        df = normalize_import_df(df_raw, ["연번", "성명", "성별", "상담여부"], header_labels)
        df.columns = [re.sub(r"\s+", "", str(col).strip()) for col in df.columns]
        rename_map = {"이름": "성명", "상담": "상담여부", "상담여부": "상담여부"}
        df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
        serial_col = "연번" if "연번" in df.columns else None
        name_col = "성명" if "성명" in df.columns else None
        gender_col = "성별" if "성별" in df.columns else None
        age_col = "만나이" if "만나이" in df.columns else ("나이" if "나이" in df.columns else None)
        consult_col = "상담여부" if "상담여부" in df.columns else None
        if not serial_col or not name_col:
            messagebox.showerror("오류", "참여자 파일에서 연번/성명(이름) 컬럼을 찾을 수 없습니다.")
            return
        mapping: Dict[str, Dict[str, str]] = {}
        for _, row in df.iterrows():
            serial = normalize_coupon_key(row.get(serial_col, ""))
            name = str(row.get(name_col, "")).strip()
            gender = str(row.get(gender_col, "")).strip() if gender_col else ""
            age = str(row.get(age_col, "")).strip() if age_col else ""
            if serial:
                mapping[serial] = {"이름": name or "", "성별": gender or "", "만나이": age or ""}
        self.participant_map = mapping
        display_df = df.copy()
        preferred_cols = [col for col in [serial_col, name_col, gender_col, age_col, consult_col] if col]
        if preferred_cols:
            display_df = display_df[preferred_cols]
        self.participant_df = display_df
        self._configure_participant_tree(list(display_df.columns))
        self._populate_participant_tree()
        self._apply_participant_names_to_import()
        self._apply_participant_names_to_ocr()
        self.status_var.set(f"참여자 로드 완료: {path.name}")

    def _apply_participant_names_to_import(self) -> None:
        if self.import_df is None or self.import_df.empty:
            return
        if COUPON_COLUMN not in self.import_df.columns or "이름" not in self.import_df.columns:
            return
        updated = False
        for idx, row in self.import_df.iterrows():
            name_raw = str(row.get("이름", "") or "").strip()
            if not name_raw.isdigit():
                continue
            coupon_key = normalize_coupon_key(name_raw)
            if not coupon_key:
                continue
            self.import_df.at[idx, COUPON_COLUMN] = name_raw
            info = self.participant_map.get(coupon_key)
            if info:
                self.import_df.at[idx, "이름"] = info.get("이름", name_raw)
                if "성별" in self.import_df.columns:
                    self.import_df.at[idx, "성별"] = info.get("성별", "")
                if "만나이" in self.import_df.columns:
                    self.import_df.at[idx, "만나이"] = info.get("만나이", "")
            updated = True
        if updated:
            self._populate_tree(self.import_tree, self.import_df)

    def _apply_participant_names_to_ocr(self) -> None:
        if self.ocr_df is None or self.ocr_df.empty or COUPON_COLUMN not in self.ocr_df.columns:
            return
        for idx, row in self.ocr_df.iterrows():
            coupon = normalize_coupon_key(row.get(COUPON_COLUMN, ""))
            if not coupon:
                continue
            info = self.participant_map.get(coupon)
            if not info:
                continue
            self.ocr_df.at[idx, "이름"] = info.get("이름", "")
            if "성별" in self.ocr_df.columns:
                self.ocr_df.at[idx, "성별"] = info.get("성별", "")
            if "만나이" in self.ocr_df.columns:
                self.ocr_df.at[idx, "만나이"] = info.get("만나이", "")
        self._populate_tree(self.ocr_tree, self.ocr_df)


    def _update_sheet_combo(self) -> None:
        if not hasattr(self, "sheet_combo"):
            return
        if self.sheet_names:
            self.sheet_combo.configure(values=self.sheet_names, state="readonly")
        else:
            self.sheet_combo.configure(values=[], state="disabled")

    def on_sheet_selected(self, _: tk.Event) -> None:
        sheet_name = self.sheet_var.get()
        if sheet_name:
            self._load_sheet_data(sheet_name)

    def _load_sheet_data(self, sheet_name: str) -> None:
        if not self._import_path:
            return
        try:
            df_raw = pd.read_excel(self._import_path, sheet_name=sheet_name, header=None)
            columns = get_sheet_columns(sheet_name)
            header_labels = get_sheet_header_labels(sheet_name)
            self.import_df = normalize_import_df(df_raw, columns, header_labels)
        except Exception as exc:
            messagebox.showerror("오류", f"시트 로드 실패: {exc}")
            return
        self.import_sheet = sheet_name
        if self.import_df is not None and sheet_name in MONTH_SHEETS:
            for col_name in ["날짜", "식당"]:
                if col_name in self.import_df.columns:
                    self.import_df[col_name] = self.import_df[col_name].replace("", np.nan)
                    self.import_df[col_name] = self.import_df[col_name].ffill().fillna("")
        if self.participant_map:
            self._apply_participant_names_to_import()
        self._configure_tree(self.import_tree, get_sheet_columns(sheet_name))
        self._populate_tree(self.import_tree, self.import_df)
        if sheet_name in MONTH_SHEETS:
            self.ocr_df = self.ocr_dfs.get(sheet_name, pd.DataFrame(columns=OCR_COLUMNS))
            self._configure_tree(self.ocr_tree, OCR_COLUMNS)
        else:
            self.ocr_df = self.ocr_dfs.get(sheet_name, pd.DataFrame(columns=get_sheet_columns(sheet_name)))
            self._configure_tree(self.ocr_tree, get_sheet_columns(sheet_name))
        self._populate_tree(self.ocr_tree, self.ocr_df)
        self._show_missing()
        if self._import_path:
            self.status_var.set(f"Excel 로드 완료: {self._import_path.name} ({sheet_name})")
    def run_analysis(self) -> None:
        if not self.photo_files:
            messagebox.showwarning("경고", "사진을 먼저 선택하세요.")
            return
        if self.import_sheet and self.import_sheet not in MONTH_SHEETS:
            messagebox.showwarning("경고", "사진 분석은 월별 시트(1월~12월)에서만 가능합니다.")
            return

        cols, rows = self.cols_var.get(), self.rows_var.get()
        all_cells: List[CouponCell] = []
        try:
            for img_path in self.photo_files:
                cells = self.processor.process_image(img_path, cols, rows)
                all_cells.extend(cells)
        except Exception as exc:
            messagebox.showerror("OCR 오류", str(exc))
            return

        records: List[Dict[str, object]] = []
        for cell in all_cells:
            rec = {col: 0 for col in AMOUNT_COLUMNS}
            rec.update(
                {
                    "날짜": "",
                    "이름": None,
                    "식당": "",
                    "사용금액": 0,
                    COUPON_COLUMN: cell.coupon_number,
                    "만나이": None,
                    "성별": None,
                }
            )
            if cell.amount in DENOMS:
                rec[f"{cell.amount}원"] = 1
                rec["사용금액"] = int(cell.amount)
            rec[OCR_TOTAL_COLUMN] = (
                int(rec.get("10000원", 0)) * 10000
                + int(rec.get("5000원", 0)) * 5000
                + int(rec.get("2000원", 0)) * 2000
                + int(rec.get("1000원", 0)) * 1000
            )
            rec["이미지"] = cell.image_path.name
            rec["위치"] = f"{cell.grid_pos[0]},{cell.grid_pos[1]}"
            rec["raw"] = " | ".join(cell.raw_texts)
            coupon_key = normalize_coupon_key(rec.get(COUPON_COLUMN, ""))
            if coupon_key and self.participant_map:
                info = self.participant_map.get(coupon_key)
                if info:
                    rec["이름"] = info.get("이름", rec["이름"])
                    rec["성별"] = info.get("성별", rec["성별"])
            records.append(rec)

        if not records:
            messagebox.showinfo("안내", "추출된 쿠폰이 없습니다.")
            return

        self.ocr_df = pd.DataFrame(records)
        self.ocr_df = self.ocr_df.reset_index(drop=True)
        # keep display columns
        display_df = self.ocr_df[OCR_COLUMNS].copy()
        display_df = display_df.fillna("")
        self._populate_tree(self.ocr_tree, display_df)
        self.status_var.set("OCR 완료")
        if self.import_sheet:
            self.ocr_dfs[self.import_sheet] = self.ocr_df
        self._show_missing()
    def on_ocr_double_click(self, event: tk.Event) -> None:
        if self.ocr_df is None:
            return
        row_id = self.ocr_tree.identify_row(event.y)
        col_id = self.ocr_tree.identify_column(event.x)
        if not row_id or not col_id:
            return
        columns = list(self.ocr_tree["columns"])
        if columns != OCR_COLUMNS:
            return
        col_index = int(col_id.replace("#", "")) - 1
        if col_index < 0 or col_index >= len(columns):
            return
        column_name = columns[col_index]
        editable = set(TABLE_COLUMNS)
        if column_name not in editable:
            return
        bbox = self.ocr_tree.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox
        value = self.ocr_tree.set(row_id, column_name)
        if column_name in AMOUNT_COLUMNS + ["사용금액"] and value == "0":
            value = ""
        if self._edit_entry is not None:
            self._edit_entry.destroy()
            self._edit_entry = None
        entry = ttk.Entry(self.ocr_tree)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, value)
        entry.focus_set()
        self._edit_entry = entry
        def save_edit(_: Optional[tk.Event] = None) -> None:
            new_value = entry.get().strip()
            if column_name in AMOUNT_COLUMNS + ["사용금액"]:
                if new_value == "":
                    new_value = "0"
                if not new_value.isdigit():
                    messagebox.showerror("오류", "금액 칸은 숫자만 입력하세요.")
                    entry.focus_set()
                    return
            row_index = self.ocr_tree.index(row_id)
            is_new_row = row_index >= len(self.ocr_df)
            if is_new_row:
                new_row = {col: 0 for col in AMOUNT_COLUMNS}
                new_row.update(
                    {
                        "날짜": "",
                        "이름": None,
                        "식당": "",
                        "사용금액": 0,
                        COUPON_COLUMN: "",
                        "만나이": None,
                        "성별": None,
                        OCR_TOTAL_COLUMN: 0,
                    }
                )
                self.ocr_df = pd.concat([self.ocr_df, pd.DataFrame([new_row])], ignore_index=True)
            if column_name == COUPON_COLUMN:
                self.ocr_df.at[row_index, column_name] = new_value
                coupon_key = normalize_coupon_key(new_value)
                if coupon_key and self.participant_map:
                    info = self.participant_map.get(coupon_key)
                    if info:
                        name_value = info.get("이름", "")
                        self.ocr_df.at[row_index, "이름"] = name_value
                        self.ocr_tree.set(row_id, "이름", name_value)
                        if "성별" in self.ocr_df.columns:
                            gender_value = info.get("성별", "")
                            self.ocr_df.at[row_index, "성별"] = gender_value
                            self.ocr_tree.set(row_id, "성별", gender_value)
                        if "만나이" in self.ocr_df.columns:
                            age_value = info.get("만나이", "")
                            self.ocr_df.at[row_index, "만나이"] = age_value
                            self.ocr_tree.set(row_id, "만나이", age_value)
            elif column_name in AMOUNT_COLUMNS:
                self.ocr_df.at[row_index, column_name] = int(new_value)
                usage = sum(int(self.ocr_df.at[row_index, col] or 0) * int(col.replace("원", "")) for col in AMOUNT_COLUMNS)
                self.ocr_df.at[row_index, "사용금액"] = usage
                self.ocr_tree.set(row_id, "사용금액", str(usage))
                self.ocr_df.at[row_index, OCR_TOTAL_COLUMN] = usage
                self.ocr_tree.set(row_id, OCR_TOTAL_COLUMN, str(usage))
            else:
                self.ocr_df.at[row_index, column_name] = new_value
            self.ocr_tree.set(row_id, column_name, new_value)
            if is_new_row:
                self._populate_tree(self.ocr_tree, self.ocr_df)
            if self.import_sheet:
                self.ocr_dfs[self.import_sheet] = self.ocr_df
            entry.destroy()
            self._edit_entry = None
            self._show_missing()
        def cancel_edit(_: Optional[tk.Event] = None) -> None:
            entry.destroy()
            self._edit_entry = None
        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", save_edit)
        entry.bind("<Escape>", cancel_edit)
    def on_ocr_single_click(self, event: tk.Event) -> None:
        if self.ocr_df is None:
            return
        columns = list(self.ocr_tree["columns"])
        if columns != OCR_COLUMNS:
            return
        row_id = self.ocr_tree.identify_row(event.y)
        col_id = self.ocr_tree.identify_column(event.x)
        if not row_id or not col_id:
            return
        col_index = int(col_id.replace("#", "")) - 1
        if col_index < 0 or col_index >= len(columns):
            return
        column_name = columns[col_index]
        if column_name != "식당":
            return
        bbox = self.ocr_tree.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox
        value = self.ocr_tree.set(row_id, column_name)
        if self._edit_entry is not None:
            self._edit_entry.destroy()
            self._edit_entry = None
        combo = ttk.Combobox(self.ocr_tree, values=self.restaurant_options, state="readonly")
        combo.place(x=x, y=y, width=w, height=h)
        if value in self.restaurant_options:
            combo.set(value)
        combo.focus_set()
        self._edit_entry = combo
        def save_select(_: Optional[tk.Event] = None) -> None:
            new_value = combo.get().strip()
            row_index = self.ocr_tree.index(row_id)
            if row_index < len(self.ocr_df):
                self.ocr_df.at[row_index, column_name] = new_value
                self.ocr_tree.set(row_id, column_name, new_value)
                if self.import_sheet:
                    self.ocr_dfs[self.import_sheet] = self.ocr_df
            combo.destroy()
            self._edit_entry = None
        def cancel_select(_: Optional[tk.Event] = None) -> None:
            combo.destroy()
            self._edit_entry = None
        combo.bind("<<ComboboxSelected>>", save_select)
        combo.bind("<FocusOut>", cancel_select)
    def on_amount_double_click(self, event: tk.Event) -> None:
        row_id = self.amount_tree.identify_row(event.y)
        col_id = self.amount_tree.identify_column(event.x)
        if not row_id or not col_id:
            return
        col_index = int(col_id.replace("#", "")) - 1
        if col_index < 0 or col_index >= len(AMOUNT_TABLE_COLUMNS):
            return
        column_name = AMOUNT_TABLE_COLUMNS[col_index]
        if column_name in ("No", "금액"):
            return
        bbox = self.amount_tree.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox
        value = self.amount_tree.set(row_id, column_name)
        if column_name in ("단가", "수량") and value == "0":
            value = ""
        if self._edit_entry is not None:
            self._edit_entry.destroy()
            self._edit_entry = None
        entry = ttk.Entry(self.amount_tree)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, value)
        entry.focus_set()
        self._edit_entry = entry
        def save_edit(_: Optional[tk.Event] = None) -> None:
            new_value = entry.get().strip()
            if column_name in ("단가", "수량"):
                if new_value == "":
                    new_value = "0"
                if not new_value.isdigit():
                    messagebox.showerror("오류", "금액/수량은 숫자로 입력하세요.")
                    entry.focus_set()
                    return
            row_index = self.amount_tree.index(row_id)
            is_new_row = row_index >= len(self.amount_df)
            if is_new_row:
                new_row = {"No": row_index + 1, "제품명": "", "단가": 0, "수량": 0, "금액": 0}
                self.amount_df = pd.concat([self.amount_df, pd.DataFrame([new_row])], ignore_index=True)
            if column_name == "제품명":
                self.amount_df.at[row_index, column_name] = new_value
            else:
                self.amount_df.at[row_index, column_name] = int(new_value)
            self._populate_amount_tree()
            entry.destroy()
            self._edit_entry = None
        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", save_edit)
    def on_ocr_right_click(self, event: tk.Event) -> None:
        row_id = self.ocr_tree.identify_row(event.y)
        if row_id:
            self.ocr_tree.selection_set(row_id)
            self.ocr_menu.tk_popup(event.x_root, event.y_root)
    def delete_selected_ocr(self) -> None:
        if self.ocr_df is None:
            return
        selected = self.ocr_tree.selection()
        if not selected:
            return
        row_id = selected[0]
        row_index = self.ocr_tree.index(row_id)
        if row_index >= len(self.ocr_df):
            return
        self.ocr_df = self.ocr_df.drop(self.ocr_df.index[row_index]).reset_index(drop=True)
        self._populate_tree(self.ocr_tree, self.ocr_df)
        if self.import_sheet:
            self.ocr_dfs[self.import_sheet] = self.ocr_df
        self._show_missing()
    def on_import_hover(self, event: tk.Event) -> None:
        row_id = self.import_tree.identify_row(event.y)
        for item in self.import_tree.get_children():
            self.import_tree.item(item, tags=())
        if row_id:
            self.import_tree.item(row_id, tags=("hover",))
    def on_import_leave(self, _: tk.Event) -> None:
        for item in self.import_tree.get_children():
            self.import_tree.item(item, tags=())
    def activate_amount_tab(self) -> None:
        self.import_label.configure(text="금액 계산")
    def _parse_target_amount(self) -> Optional[int]:
        raw = self.amount_target_var.get().strip()
        if raw == "":
            messagebox.showwarning("경고", "원하는 금액을 입력하세요.")
            return None
        if not raw.isdigit():
            messagebox.showwarning("경고", "원하는 금액은 숫자로 입력하세요.")
            return None
        return int(raw)
    def _solve_quantities(self, prices: List[int], target: int, min_each: int = 0) -> Optional[List[int]]:
        if target < 0 or not prices:
            return None
        base_total = sum(price * min_each for price in prices)
        if target < base_total:
            return None
        remaining_target = target - base_total
        denom = reduce(gcd, prices)
        if remaining_target % denom != 0:
            return None
        attempts = 2000
        for _ in range(attempts):
            remaining = remaining_target
            qtys = [min_each] * len(prices)
            for idx in range(len(prices) - 1):
                price = prices[idx]
                max_qty = remaining // price
                qty = random.randint(0, max_qty) if max_qty > 0 else 0
                qtys[idx] += qty
                remaining -= qty * price
            last_price = prices[-1]
            if remaining % last_price == 0:
                qtys[-1] += remaining // last_price
                return qtys
        return None
    def adjust_amounts(self) -> None:
        target = self._parse_target_amount()
        if target is None:
            return
        if self.amount_df is None or self.amount_df.empty:
            messagebox.showinfo("안내", "제품 목록을 먼저 입력하세요.")
            return
        eligible: List[Tuple[int, int]] = []
        for idx, row in self.amount_df.iterrows():
            name = str(row.get("제품명", "")).strip()
            price = int(row.get("단가", 0) or 0)
            if name and price <= 0:
                messagebox.showwarning("경고", "제품명에 해당하는 단가가 0입니다. 단가를 입력하세요.")
                return
            if name and price > 0:
                eligible.append((idx, price))
        if not eligible:
            messagebox.showinfo("안내", "제품명과 금액이 입력된 항목이 없습니다.")
            return

        selected_indices = [idx for idx, _ in eligible]
        prices = [price for _, price in eligible]
        qtys = self._solve_quantities(prices, target, min_each=1)
        if qtys is None:
            messagebox.showwarning("경고", "모든 제품 수량을 1개 이상으로 설정할 수 없습니다.")
            return

        self.amount_df["수량"] = 0
        for idx, qty in zip(selected_indices, qtys):
            self.amount_df.at[idx, "수량"] = int(qty)
        self._populate_amount_tree()
    def export_amount_table(self) -> None:
        if self.amount_df is None or self.amount_df.empty:
            messagebox.showinfo("안내", "내보낼 데이터가 없습니다.")
            return
        save_path = filedialog.asksaveasfilename(
            parent=self.root,
            title="금액 내보내기",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not save_path:
            return
        export_df = self.amount_df.copy()
        export_df = export_df[
            (export_df["제품명"].astype(str).str.strip() != "")
            | (export_df["단가"].fillna(0).astype(int) > 0)
            | (export_df["수량"].fillna(0).astype(int) > 0)
        ].reset_index(drop=True)
        export_df["No"] = [idx + 1 for idx in range(len(export_df))]
        export_df["금액"] = (export_df["단가"].fillna(0).astype(int)
                           * export_df["수량"].fillna(0).astype(int))
        total_amount = int(export_df["금액"].fillna(0).astype(int).sum())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(AMOUNT_TABLE_COLUMNS)
        for _, row in export_df.iterrows():
            ws.append([row.get(col, "") for col in AMOUNT_TABLE_COLUMNS])
        ws.append(["", "", "Total", "", total_amount])

        thin = openpyxl.styles.Side(style="thin", color="000000")
        border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(AMOUNT_TABLE_COLUMNS)):
            for cell in row:
                cell.border = border

        wb.save(save_path)
        self._last_export_path = Path(save_path)
        self._save_last_export_path()
        messagebox.showinfo("완료", f"저장 완료: {Path(save_path).name}")
    def open_export_folder(self) -> None:
        if not self._last_export_path:
            messagebox.showinfo("안내", "Export된 파일이 없습니다.")
            return
        folder = self._last_export_path
        if not folder.is_dir():
            folder = folder.parent
        if folder.exists():
            os.startfile(str(folder))
        else:
            messagebox.showinfo("안내", "저장 폴더를 찾을 수 없습니다.")
    def _show_missing(self) -> None:
        self.missing_text.delete("1.0", tk.END)
        if self.import_df is None or self.ocr_df is None:
            return
        if COUPON_COLUMN not in self.import_df.columns or COUPON_COLUMN not in self.ocr_df.columns:
            return
        base_numbers = set(self.import_df[COUPON_COLUMN].dropna().astype(str))
        missing_rows = []
        for _, row in self.ocr_df.iterrows():
            num = str(row[COUPON_COLUMN]) if pd.notna(row[COUPON_COLUMN]) else None
            if not num or num not in base_numbers:
                missing_rows.append(
                    f"{row.get('이미지','')}, 위치 {row.get('위치','')}, 쿠폰번호: {num or '미인식'}"
                )
        if missing_rows:
            self.missing_text.insert(tk.END, "\n".join(missing_rows))
        else:
            self.missing_text.insert(tk.END, "모든 쿠폰번호가 Import 파일에 존재합니다.")
    def export_excel(self) -> None:
        if self.import_df is None:
            messagebox.showwarning("경고", "먼저 Excel을 Import 하세요.")
            return
        if self.ocr_df is None:
            messagebox.showwarning("경고", "먼저 OCR Analysis를 실행하세요.")
            return
        if self._export_in_progress:
            messagebox.showinfo("안내", "Export가 진행 중입니다. 잠시만 기다려 주세요.")
            return

        save_path = filedialog.asksaveasfilename(
            parent=self.root,
            title="Export 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not save_path:
            return
        self._export_in_progress = True
        self.status_var.set("Export 준비 중...")
        self.root.update_idletasks()

        def worker() -> None:
            error_msg = None
            try:
                self.root.after(0, lambda: self.status_var.set("Export 계산 중..."))
                if self.import_sheet in MONTH_SHEETS:
                    merged = self.merge_data()
                else:
                    merged = self.import_df.copy() if self.import_df is not None else pd.DataFrame()
                target_path = Path(save_path)
                temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
                log_path = Path("export_error.log")
                log_path.write_text(
                    f"start export: {target_path}\n",
                    encoding="utf-8",
                )
                self.root.after(0, lambda: self.status_var.set("Export 저장 중..."))
                build_export_workbook(
                    merged,
                    temp_path,
                    template_path=self._import_path,
                    original_df=self.import_df,
                    sheet_name=self.import_sheet,
                )
                log_path.write_text(
                    f"saved temp: {temp_path}\n",
                    encoding="utf-8",
                    errors="replace",
                )
                if not temp_path.exists() or temp_path.stat().st_size == 0:
                    raise RuntimeError("임시 파일 생성에 실패했습니다.")
                temp_path.replace(target_path)
                log_path.write_text(
                    "replaced temp -> target\n",
                    encoding="utf-8",
                    errors="replace",
                )
            except PermissionError:
                error_msg = "파일이 열려 있어 저장할 수 없습니다. Excel에서 닫고 다시 시도하세요."
            except Exception as exc:
                log_path.write_text(traceback.format_exc(), encoding="utf-8")
                error_msg = f"Export 실패: {exc} (export_error.log 확인)"

            def on_done() -> None:
                self._export_in_progress = False
                if error_msg:
                    messagebox.showerror("오류", error_msg)
                    self.status_var.set("Export 실패")
                    return
                self._last_export_path = Path(save_path)
                self._save_last_export_path()
                self.status_var.set(f"Export 완료: {Path(save_path).name}")
                messagebox.showinfo("완료", "Export 완료되었습니다.")

            self.root.after(0, on_done)

        threading.Thread(target=worker, daemon=True).start()
    def merge_data(self) -> pd.DataFrame:
        base = self.import_df.copy() if self.import_df is not None else pd.DataFrame(columns=TABLE_COLUMNS)
        if COUPON_COLUMN not in base.columns:
            return base
        if self.participant_map and "이름" in base.columns:
            for idx, row in base.iterrows():
                name_raw = str(row.get("이름", "") or "").strip()
                if not name_raw.isdigit():
                    continue
                coupon_key = normalize_coupon_key(name_raw)
                if not coupon_key:
                    continue
                base.at[idx, COUPON_COLUMN] = name_raw
                info = self.participant_map.get(coupon_key)
                if info:
                    base.at[idx, "이름"] = info.get("이름", name_raw)
                    if "성별" in base.columns:
                        base.at[idx, "성별"] = info.get("성별", "")
                    if "만나이" in base.columns:
                        base.at[idx, "만나이"] = info.get("만나이", "")
        if "날짜" in base.columns:
            base["날짜"] = base["날짜"].replace("", np.nan)
            base["날짜"] = base["날짜"].ffill().fillna("")
        ocr_df = self.ocr_df.copy()
        if ocr_df is None or ocr_df.empty:
            return base
        if "날짜" in ocr_df.columns:
            today_str = datetime.now().strftime("%Y.%m.%d")
            ocr_df["날짜"] = ocr_df["날짜"].replace("", np.nan)
            ocr_df["날짜"] = ocr_df["날짜"].fillna(today_str)
        ocr_df[COUPON_COLUMN] = ocr_df[COUPON_COLUMN].fillna("").astype(str)
        if "식당" in ocr_df.columns:
            ocr_df["식당"] = ocr_df["식당"].replace("", np.nan)
            filled_mask = ocr_df["식당"].isna()
            ocr_df["식당"] = ocr_df["식당"].ffill().fillna("")
            ocr_df["__restaurant_filled"] = np.where(filled_mask & (ocr_df["식당"] != ""), 1, 0)
        else:
            ocr_df["__restaurant_filled"] = 0
        meta_cols = ["날짜", "이름", "식당", "만나이", "성별"]
        ocr_meta = (
            ocr_df.groupby(COUPON_COLUMN)[meta_cols]
            .agg(lambda x: next((v for v in x if pd.notna(v) and str(v).strip() != ""), ""))
            .reset_index()
        )
        merged = base.copy()
        for col in ["날짜", "이름", "식당", "만나이", "성별"]:
            if col not in merged.columns:
                merged[col] = ""
        base_usage = merged["사용금액"].fillna(0).astype(int) if "사용금액" in merged.columns else pd.Series([0] * len(merged))
        denom_usage = (
            merged.get("10000원", 0).fillna(0).astype(int) * 10000
            + merged.get("5000원", 0).fillna(0).astype(int) * 5000
            + merged.get("2000원", 0).fillna(0).astype(int) * 2000
            + merged.get("1000원", 0).fillna(0).astype(int) * 1000
        )
        merged["사용금액"] = np.where(base_usage > 0, base_usage, denom_usage)
        if {"날짜", "이름", "식당"}.issubset(ocr_df.columns):
            ocr_df[OCR_TOTAL_COLUMN] = (
                ocr_df.get("10000원", 0).fillna(0).astype(int) * 10000
                + ocr_df.get("5000원", 0).fillna(0).astype(int) * 5000
                + ocr_df.get("2000원", 0).fillna(0).astype(int) * 2000
                + ocr_df.get("1000원", 0).fillna(0).astype(int) * 1000
            )
            ocr_key_meta = (
                ocr_df.groupby(["날짜", "이름", "식당"])[[COUPON_COLUMN, "성별", "만나이"]]
                .agg(lambda x: next((v for v in x if pd.notna(v) and str(v).strip() != ""), ""))
                .reset_index()
            )
            ocr_match = (
                ocr_df.groupby(["날짜", "이름", "식당"])[AMOUNT_COLUMNS + [OCR_TOTAL_COLUMN, "__restaurant_filled"]]
                .sum()
                .reset_index()
            )
            ocr_match = ocr_match.merge(ocr_key_meta, on=["날짜", "이름", "식당"], how="left")
            merged = merged.merge(
                ocr_match,
                on=["날짜", "이름", "식당"],
                how="left",
                suffixes=("", "_match"),
            )
            match_total = merged[OCR_TOTAL_COLUMN].fillna(0).astype(int)
            merged["__restaurant_filled"] = merged["__restaurant_filled"].fillna(0).astype(int)
            for col in AMOUNT_COLUMNS:
                base_col = f"{col}_base"
                base_vals = merged[base_col].fillna(0).astype(int) if base_col in merged.columns else merged[col].fillna(0).astype(int)
                merged[col] = np.where(
                    match_total > 0,
                    base_vals + merged[f"{col}_match"].fillna(0).astype(int),
                    merged[col],
                )
            merged["사용금액"] = np.where(
                match_total > 0,
                base_usage + match_total,
                base_usage,
            )
            merged["__amounts_changed"] = np.where(match_total > 0, 1, 0)
            merged = merged.drop(columns=[col for col in [OCR_TOTAL_COLUMN] + [f"{c}_match" for c in AMOUNT_COLUMNS] if col in merged.columns])
            base_keys = merged[["날짜", "이름", "식당"]].astype(str)
            base_key_set = set(tuple(row) for row in base_keys.values.tolist())
            new_rows: List[Dict[str, object]] = []
            for _, row in ocr_match.iterrows():
                key = (str(row["날짜"]), str(row["이름"]), str(row["식당"]))
                if key in base_key_set:
                    continue
                new_row = {col: "" for col in TABLE_COLUMNS}
                new_row["날짜"] = row["날짜"]
                new_row["이름"] = row["이름"]
                new_row["식당"] = row["식당"]
                new_row[COUPON_COLUMN] = row.get(COUPON_COLUMN, "")
                new_row["성별"] = row.get("성별", "")
                new_row["만나이"] = row.get("만나이", "")
                for col in AMOUNT_COLUMNS:
                    new_row[col] = int(row.get(col, 0))
                new_row["사용금액"] = int(row.get(OCR_TOTAL_COLUMN, 0))
                new_row["__restaurant_filled"] = int(row.get("__restaurant_filled", 0))
                new_row["__amounts_changed"] = 1
                new_rows.append(new_row)
            if new_rows:
                merged = pd.concat([merged, pd.DataFrame(new_rows)], ignore_index=True)
        name_map: Dict[str, str] = {}
        for _, row in merged.iterrows():
            coupon = normalize_coupon_key(row.get(COUPON_COLUMN, ""))
            name = str(row.get("이름", "")).strip()
            if coupon and name and not name.isdigit():
                name_map[coupon] = name

        def resolve_numeric_name(row: pd.Series) -> Tuple[str, str]:
            name_raw = str(row.get("이름", "") or "").strip()
            coupon_raw = str(row.get(COUPON_COLUMN, "") or "").strip()
            if name_raw.isdigit():
                coupon_key = normalize_coupon_key(name_raw)
                if coupon_key:
                    coupon_raw = name_raw
                    if self.participant_map:
                        info = self.participant_map.get(coupon_key)
                        if info:
                            return coupon_raw, info.get("이름", name_raw)
                    return coupon_raw, name_map.get(coupon_key, name_raw)
            return coupon_raw, name_raw

        if "이름" in merged.columns and COUPON_COLUMN in merged.columns:
            updated = merged.apply(resolve_numeric_name, axis=1, result_type="expand")
            merged[COUPON_COLUMN] = updated[0]
            merged["이름"] = updated[1]
        keep_cols = TABLE_COLUMNS
        for col in keep_cols:
            if col not in merged.columns:
                merged[col] = ""
        extra_cols = [col for col in ["__restaurant_filled", "__amounts_changed"] if col in merged.columns]
        merged = merged[keep_cols + extra_cols]
        merged[["이름", "만나이", "성별", "날짜", "식당"]] = merged[["이름", "만나이", "성별", "날짜", "식당"]].fillna("")
        if "날짜" in merged.columns:
            merged["_date_sort"] = pd.to_datetime(merged["날짜"], errors="coerce", format="%Y.%m.%d")
            merged = merged.sort_values(by="_date_sort", ascending=False, na_position="last").drop(columns=["_date_sort"])
        if "__restaurant_filled" in merged.columns:
            merged["__restaurant_filled"] = merged["__restaurant_filled"].fillna(0).astype(int)
        if "__amounts_changed" in merged.columns:
            merged["__amounts_changed"] = merged["__amounts_changed"].fillna(0).astype(int)
        return merged
def main() -> None:
    root = tk.Tk()
    App(root)
    root.mainloop()
if __name__ == "__main__":
    main()
